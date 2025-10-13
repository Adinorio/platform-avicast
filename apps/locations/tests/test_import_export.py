"""
Test cases for Census Data Import/Export functionality
AGENTS.md §6.1 Testing - Comprehensive test coverage for import validation

Run tests with:
    python manage.py test apps.locations.tests.test_import_export
"""

from datetime import date
from io import BytesIO
from django.test import TestCase, Client
from django.contrib.auth import get_user_model
from openpyxl import Workbook

from apps.locations.models import Site, CensusYear, CensusMonth, Census, CensusObservation
from apps.fauna.models import Species
from apps.locations.utils.excel_handler import CensusExcelHandler, ExcelImportError


User = get_user_model()


class ExcelTemplateTestCase(TestCase):
    """Test Excel template generation"""
    
    def test_generate_template(self):
        """Test that template generation creates valid workbook"""
        wb = CensusExcelHandler.generate_template()
        self.assertIsNotNone(wb)
        
        # Check sheets exist
        self.assertIn("Census Data", wb.sheetnames)
        self.assertIn("Instructions", wb.sheetnames)
        
        # Check headers
        ws = wb["Census Data"]
        headers = [cell.value for cell in ws[1]]
        expected_headers = CensusExcelHandler.HEADER_ROW
        self.assertEqual(headers, expected_headers)


class ExcelImportValidationTestCase(TestCase):
    """Test Excel import validation logic"""
    
    def setUp(self):
        """Set up test data"""
        self.user = User.objects.create_user(
            employee_id="TEST001",
            username="testuser",
            password="testpass123"
        )
        
        self.site = Site.objects.create(
            name="Test Site",
            site_type="WETLAND",
            status="active",
            created_by=self.user
        )
        
        self.species = Species.objects.create(
            name="Test Bird",
            scientific_name="Testus birdus",
            iucn_status="LC"
        )
    
    def test_validate_row_success(self):
        """Test validation of valid row data"""
        row_data = {
            'site_name': 'Test Site',
            'year': 2024,
            'month': 1,
            'census_date': '2024-01-15',
            'species_name': 'Test Bird',
            'count': 10,
        }
        
        is_valid, error = CensusExcelHandler.validate_row(row_data, 2)
        self.assertTrue(is_valid)
        self.assertEqual(error, "")
    
    def test_validate_row_missing_required_field(self):
        """Test validation fails for missing required fields"""
        row_data = {
            'site_name': 'Test Site',
            'year': 2024,
            'month': 1,
            # Missing census_date
            'species_name': 'Test Bird',
            'count': 10,
        }
        
        is_valid, error = CensusExcelHandler.validate_row(row_data, 2)
        self.assertFalse(is_valid)
        self.assertIn("Missing required field", error)
    
    def test_validate_row_invalid_year(self):
        """Test validation fails for invalid year"""
        row_data = {
            'site_name': 'Test Site',
            'year': 1800,  # Too old
            'month': 1,
            'census_date': '2024-01-15',
            'species_name': 'Test Bird',
            'count': 10,
        }
        
        is_valid, error = CensusExcelHandler.validate_row(row_data, 2)
        self.assertFalse(is_valid)
        self.assertIn("Invalid year", error)
    
    def test_validate_row_invalid_month(self):
        """Test validation fails for invalid month"""
        row_data = {
            'site_name': 'Test Site',
            'year': 2024,
            'month': 13,  # Invalid month
            'census_date': '2024-01-15',
            'species_name': 'Test Bird',
            'count': 10,
        }
        
        is_valid, error = CensusExcelHandler.validate_row(row_data, 2)
        self.assertFalse(is_valid)
        self.assertIn("Month must be between 1-12", error)
    
    def test_validate_row_invalid_count(self):
        """Test validation fails for invalid count"""
        row_data = {
            'site_name': 'Test Site',
            'year': 2024,
            'month': 1,
            'census_date': '2024-01-15',
            'species_name': 'Test Bird',
            'count': 0,  # Must be positive
        }
        
        is_valid, error = CensusExcelHandler.validate_row(row_data, 2)
        self.assertFalse(is_valid)
        self.assertIn("Count must be a positive number", error)
    
    def test_validate_row_nonexistent_site(self):
        """Test validation fails for nonexistent site"""
        row_data = {
            'site_name': 'Nonexistent Site',
            'year': 2024,
            'month': 1,
            'census_date': '2024-01-15',
            'species_name': 'Test Bird',
            'count': 10,
        }
        
        is_valid, error = CensusExcelHandler.validate_row(row_data, 2)
        self.assertFalse(is_valid)
        self.assertIn("Site", error)
        self.assertIn("not found", error)


class ExcelImportIntegrationTestCase(TestCase):
    """Test full Excel import workflow"""
    
    def setUp(self):
        """Set up test data"""
        self.user = User.objects.create_user(
            employee_id="TEST001",
            username="testuser",
            password="testpass123"
        )
        
        self.site = Site.objects.create(
            name="Test Site",
            site_type="WETLAND",
            status="active",
            created_by=self.user
        )
        
        self.species1 = Species.objects.create(
            name="Chinese Egret",
            scientific_name="Egretta eulophotes",
            iucn_status="VU"
        )
        
        self.species2 = Species.objects.create(
            name="Black-faced Spoonbill",
            scientific_name="Platalea minor",
            iucn_status="EN"
        )
    
    def create_test_excel(self, data_rows):
        """Helper to create test Excel file"""
        wb = Workbook()
        ws = wb.active
        
        # Add headers
        ws.append(CensusExcelHandler.HEADER_ROW)
        
        # Add data rows
        for row in data_rows:
            ws.append(row)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    def test_import_single_observation(self):
        """Test importing a single observation"""
        data_rows = [
            ["Test Site", 2024, 1, "2024-01-15", "Chinese Egret", 45, "Sunny", "Test note"]
        ]
        
        excel_file = self.create_test_excel(data_rows)
        results = CensusExcelHandler.import_from_excel(excel_file, self.user)
        
        self.assertEqual(results['successful'], 1)
        self.assertEqual(results['skipped'], 0)
        self.assertEqual(results['created_census'], 1)
        self.assertEqual(results['created_observations'], 1)
        
        # Verify data was created
        self.assertEqual(Census.objects.count(), 1)
        self.assertEqual(CensusObservation.objects.count(), 1)
        
        obs = CensusObservation.objects.first()
        self.assertEqual(obs.species, self.species1)
        self.assertEqual(obs.count, 45)
    
    def test_import_multiple_species_same_census(self):
        """Test importing multiple species for same census date"""
        data_rows = [
            ["Test Site", 2024, 1, "2024-01-15", "Chinese Egret", 45, "Sunny", "Morning"],
            ["Test Site", 2024, 1, "2024-01-15", "Black-faced Spoonbill", 23, "Sunny", "Morning"]
        ]
        
        excel_file = self.create_test_excel(data_rows)
        results = CensusExcelHandler.import_from_excel(excel_file, self.user)
        
        self.assertEqual(results['successful'], 2)
        self.assertEqual(results['created_census'], 1)  # Only one census
        self.assertEqual(results['created_observations'], 2)  # Two observations
        
        census = Census.objects.first()
        self.assertEqual(census.observations.count(), 2)
    
    def test_import_invalid_species(self):
        """Test that invalid species are skipped"""
        data_rows = [
            ["Test Site", 2024, 1, "2024-01-15", "Nonexistent Bird", 10, "", ""]
        ]
        
        excel_file = self.create_test_excel(data_rows)
        results = CensusExcelHandler.import_from_excel(excel_file, self.user)
        
        self.assertEqual(results['successful'], 0)
        self.assertEqual(results['skipped'], 1)
        self.assertGreater(len(results['errors']), 0)
        
        # No data should be created
        self.assertEqual(Census.objects.count(), 0)
        self.assertEqual(CensusObservation.objects.count(), 0)
    
    def test_import_duplicate_observation(self):
        """Test that duplicate observations are handled correctly"""
        # First import
        data_rows = [
            ["Test Site", 2024, 1, "2024-01-15", "Chinese Egret", 45, "Sunny", ""]
        ]
        excel_file = self.create_test_excel(data_rows)
        CensusExcelHandler.import_from_excel(excel_file, self.user)
        
        # Second import with same data
        excel_file = self.create_test_excel(data_rows)
        results = CensusExcelHandler.import_from_excel(excel_file, self.user)
        
        # Should merge counts
        obs = CensusObservation.objects.first()
        self.assertEqual(obs.count, 90)  # 45 + 45


class ExcelExportTestCase(TestCase):
    """Test Excel export functionality"""
    
    def setUp(self):
        """Set up test data"""
        self.user = User.objects.create_user(
            employee_id="TEST001",
            username="testuser",
            password="testpass123"
        )
        
        self.site = Site.objects.create(
            name="Test Site",
            site_type="WETLAND",
            status="active",
            created_by=self.user
        )
        
        self.species = Species.objects.create(
            name="Chinese Egret",
            scientific_name="Egretta eulophotes",
            iucn_status="VU"
        )
        
        # Create test census data
        year = CensusYear.objects.create(site=self.site, year=2024)
        month = CensusMonth.objects.create(year=year, month=1)
        census = Census.objects.create(
            month=month,
            census_date=date(2024, 1, 15),
            lead_observer=self.user
        )
        CensusObservation.objects.create(
            census=census,
            species=self.species,
            species_name=self.species.name,
            count=45
        )
    
    def test_export_all_data(self):
        """Test exporting all census data"""
        wb = CensusExcelHandler.export_to_excel()
        ws = wb.active
        
        # Check headers
        headers = [cell.value for cell in ws[1]]
        self.assertEqual(headers, CensusExcelHandler.HEADER_ROW)
        
        # Check data row exists
        self.assertIsNotNone(ws.cell(row=2, column=1).value)
    
    def test_export_with_site_filter(self):
        """Test exporting with site filter"""
        filters = {'site_id': self.site.id}
        wb = CensusExcelHandler.export_to_excel(filters)
        
        self.assertIsNotNone(wb)


class ImportExportViewsTestCase(TestCase):
    """Test import/export views"""
    
    def setUp(self):
        """Set up test client and user"""
        self.client = Client()
        self.user = User.objects.create_user(
            employee_id="TEST001",
            username="testuser",
            password="testpass123"
        )
        self.client.login(username="testuser", password="testpass123")
    
    def test_import_export_hub_view(self):
        """Test that import/export hub loads"""
        response = self.client.get('/locations/data/')
        self.assertEqual(response.status_code, 200)
    
    def test_download_template_view(self):
        """Test template download"""
        response = self.client.get('/locations/data/template/')
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    def test_import_page_loads(self):
        """Test import page loads"""
        response = self.client.get('/locations/data/import/')
        self.assertEqual(response.status_code, 200)
    
    def test_export_page_loads(self):
        """Test export page loads"""
        response = self.client.get('/locations/data/export/')
        self.assertEqual(response.status_code, 200)
    
    def test_census_totals_view_loads(self):
        """Test census totals view loads"""
        response = self.client.get('/locations/data/totals/')
        self.assertEqual(response.status_code, 200)


# TODO: Add more test cases
# - Test large file imports (100+ rows)
# - Test concurrent imports
# - Test export with complex filters
# - Test species breakdown view
# - Test error handling for corrupted Excel files
# - Test permission-based access (different user roles)

