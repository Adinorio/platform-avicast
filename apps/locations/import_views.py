"""
Import/export views for locations app
"""

import csv
import io
import json

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook, load_workbook

from django.contrib.admin.views.decorators import staff_member_required as staff_required
from .forms import CensusImportForm
from .models import CensusObservation, MobileDataImport, Site, SpeciesObservation, BulkCensusImport, log_census_activity


@login_required
@staff_required
def census_import(request, site_id):
    """Import census data from CSV/Excel file"""
    site = get_object_or_404(Site, id=site_id)

    if request.method == "POST":
        form = CensusImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Create census observation
                    census = CensusObservation.objects.create(
                        site=site,
                        observation_date=form.cleaned_data["observation_date"],
                        observer=request.user,
                        weather_conditions=form.cleaned_data["weather_conditions"],
                        notes=form.cleaned_data["notes"],
                    )

                    # Process the uploaded file
                    file = form.cleaned_data["file"]
                    if file.name.endswith(".csv"):
                        success_count = _process_csv_import(file, census)
                    else:
                        success_count = _process_excel_import(file, census)

                    messages.success(
                        request, f"Successfully imported {success_count} species observations!"
                    )
                    return redirect("locations:site_detail", site_id=site.id)

            except Exception as e:
                messages.error(request, f"Import failed: {str(e)}")
    else:
        form = CensusImportForm()

    context = {"form": form, "site": site}
    return render(request, "locations/census_import.html", context)


def _process_csv_import(file, census):
    """Process CSV file import"""
    decoded_file = file.read().decode("utf-8")
    csv_data = csv.DictReader(io.StringIO(decoded_file))

    success_count = 0
    for row in csv_data:
        try:
            SpeciesObservation.objects.create(
                census=census,
                species_name=row.get("species_name", "").strip(),
                count=int(row.get("count", 1)),
                behavior_notes=row.get("behavior_notes", "").strip(),
            )
            success_count += 1
        except (ValueError, KeyError):
            continue

    return success_count


def _process_excel_import(file, census):
    """Process Excel file import"""
    workbook = load_workbook(file)
    sheet = workbook.active

    success_count = 0
    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2):
        try:
            row_data = {headers[i]: cell.value for i, cell in enumerate(row)}

            if row_data.get("species_name") and row_data.get("count"):
                SpeciesObservation.objects.create(
                    census=census,
                    species_name=str(row_data["species_name"]).strip(),
                    count=int(row_data["count"]),
                    behavior_notes=str(row_data.get("behavior_notes", "")).strip(),
                )
                success_count += 1
        except (ValueError, KeyError):
            continue

    return success_count


@login_required
@staff_required
def census_export_csv(request, site_id):
    """Export census data as CSV"""
    site = get_object_or_404(Site, id=site_id)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{site.name}_census_data.csv"'

    writer = csv.writer(response)
    writer.writerow(["Date", "Species", "Count", "Behavior Notes", "Weather", "Observer"])

    census_observations = CensusObservation.objects.filter(site=site).order_by("-observation_date")

    for census in census_observations:
        for species_obs in census.species_observations.all():
            writer.writerow(
                [
                    census.observation_date,
                    species_obs.species_name,
                    species_obs.count,
                    species_obs.behavior_notes,
                    census.weather_conditions,
                    census.observer.username if census.observer else "",
                ]
            )

    return response


@login_required
@staff_required
def census_export_excel(request, site_id):
    """Export census data as Excel"""
    site = get_object_or_404(Site, id=site_id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{site.name} Census Data"

    # Headers
    headers = ["Date", "Species", "Count", "Behavior Notes", "Weather", "Observer"]
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)

    # Data
    row = 2
    census_observations = CensusObservation.objects.filter(site=site).order_by("-observation_date")

    for census in census_observations:
        for species_obs in census.species_observations.all():
            sheet.cell(row=row, column=1, value=census.observation_date)
            sheet.cell(row=row, column=2, value=species_obs.species_name)
            sheet.cell(row=row, column=3, value=species_obs.count)
            sheet.cell(row=row, column=4, value=species_obs.behavior_notes)
            sheet.cell(row=row, column=5, value=census.weather_conditions)
            sheet.cell(row=row, column=6, value=census.observer.username if census.observer else "")
            row += 1

    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (AttributeError, TypeError):
                pass

        # Mobile Data Import Views


@login_required
@staff_required
def mobile_import_list(request):
    """List all mobile data import requests with comprehensive statistics"""
    if request.user.role in ['ADMIN', 'SUPERADMIN']:
        imports = MobileDataImport.objects.all()
    else:
        imports = MobileDataImport.objects.filter(submitted_by=request.user)

    # Calculate statistics before filtering
    stats = {
        'total': imports.count(),
        'pending': imports.filter(status='pending').count(),
        'approved': imports.filter(status='approved').count(),
        'rejected': imports.filter(status='rejected').count(),
        'processed': imports.filter(status='processed').count(),
    }

    # Filter by status
    status_filter = request.GET.get("status", "")
    if status_filter:
        imports = imports.filter(status=status_filter)

    # Filter by site
    site_filter = request.GET.get("site", "")
    if site_filter:
        imports = imports.filter(site_id=site_filter)

    # Order by most recent first
    imports = imports.select_related('site', 'submitted_by', 'reviewed_by').order_by('-created_at')

    from django.core.paginator import Paginator
    paginator = Paginator(imports, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Get all sites for filter
    sites = Site.objects.all().order_by('name')

    context = {
        "page_obj": page_obj,
        "status_filter": status_filter,
        "site_filter": site_filter,
        "status_choices": MobileDataImport.IMPORT_STATUS,
        "stats": stats,
        "sites": sites,
        "is_admin": request.user.role in ['ADMIN', 'SUPERADMIN'],
    }

    return render(request, "locations/mobile_import_list.html", context)


@login_required
@staff_required
def submit_mobile_import(request):
    """Submit a new mobile data import request"""
    if request.method == "POST":
        try:
            import_data = json.loads(request.body)
            site_id = import_data.get("site_id")

            site = get_object_or_404(Site, id=site_id)

            # Create the import request
            mobile_import = MobileDataImport.objects.create(
                import_data=import_data, site=site, submitted_by=request.user, status="pending"
            )

            return JsonResponse(
                {
                    "success": True,
                    "message": "Import request submitted successfully",
                    "import_id": str(mobile_import.id),
                }
            )

        except Exception as e:
            return JsonResponse({"success": False, "message": f"Error submitting import: {str(e)}"})

    # GET request - show form
    sites = Site.objects.all()
    return render(request, "locations/submit_mobile_import.html", {"sites": sites})


@login_required
@staff_required
def review_mobile_import(request, import_id):
    """Review and approve/reject mobile data import requests with detailed preview"""
    mobile_import = get_object_or_404(MobileDataImport, id=import_id)

    # Parse import data for preview
    import_data = mobile_import.import_data
    species_data = import_data.get('species', [])

    # Calculate totals for preview
    total_species = len(species_data)
    total_birds = sum(species.get('count', 0) for species in species_data)

    # Check for existing census data for comparison
    observation_date = import_data.get('observation_date')
    existing_census = None
    existing_species_count = 0
    existing_birds_count = 0
    conflicts = []
    new_species = []

    if observation_date:
        existing_census = CensusObservation.objects.filter(
            site=mobile_import.site,
            observation_date=observation_date
        ).first()

        if existing_census:
            existing_species_count = existing_census.get_total_species_count()
            existing_birds_count = existing_census.get_total_birds_count()

            # Compare imported vs existing data
            for imported_species in species_data:
                species_name = imported_species.get('species_name', '')
                imported_count = imported_species.get('count', 0)

                # Find existing species observation
                existing_species_obs = SpeciesObservation.objects.filter(
                    census=existing_census,
                    species_name=species_name
                ).first()

                if existing_species_obs:
                    existing_count = existing_species_obs.count
                    if existing_count != imported_count:
                        conflicts.append({
                            'species_name': species_name,
                            'existing_count': existing_count,
                            'imported_count': imported_count,
                            'difference': imported_count - existing_count
                        })
                else:
                    new_species.append(imported_species)

        # Mark species data with conflict status
        for species in species_data:
            species_name = species.get('species_name', '')
            if any(conflict.get('species_name') == species_name for conflict in conflicts):
                species['conflict'] = True
            else:
                species['conflict'] = False

    if request.method == "POST":
        action = request.POST.get("action")
        notes = request.POST.get("review_notes", "")

        try:
            if action == "approve":
                mobile_import.approve(request.user)
                messages.success(
                    request,
                    f"Import request approved. {total_species} species with {total_birds} birds ready to process."
                )
            elif action == "reject":
                mobile_import.reject(request.user, notes)
                messages.warning(request, "Import request rejected.")
            elif action == "process":
                if mobile_import.status != "approved":
                    messages.error(request, "Only approved imports can be processed.")
                else:
                    census = mobile_import.process_import()
                    messages.success(
                        request,
                        f"Import processed successfully! Census observation created with {total_species} species."
                    )
                    return redirect("locations:site_detail", site_id=mobile_import.site.id)

            return redirect("locations:mobile_import_list")

        except Exception as e:
            messages.error(request, f"Error processing request: {str(e)}")

    context = {
        "mobile_import": mobile_import,
        "import_data_json": json.dumps(import_data, indent=2),
        "species_data": species_data,
        "total_species": total_species,
        "total_birds": total_birds,
        "observation_date": import_data.get('observation_date', 'Not specified'),
        "weather_conditions": import_data.get('weather_conditions', 'Not specified'),
        "notes": import_data.get('notes', ''),
        "is_admin": request.user.role in ['ADMIN', 'SUPERADMIN'],
        "can_edit": mobile_import.status == 'pending' and request.user.role in ['ADMIN', 'SUPERADMIN'],
        # Comparison data
        "existing_census": existing_census,
        "existing_species_count": existing_species_count,
        "existing_birds_count": existing_birds_count,
        "conflicts": conflicts,
        "new_species": new_species,
    }

    return render(request, "locations/review_mobile_import.html", context)


@login_required
@staff_required
def bulk_import_actions(request):
    """Handle bulk actions on mobile imports"""
    if request.method == "POST":
        import_ids = request.POST.getlist("import_ids")
        action = request.POST.get("bulk_action")

        imports = MobileDataImport.objects.filter(id__in=import_ids)
        processed = 0

        for mobile_import in imports:
            try:
                if action == "approve":
                    mobile_import.approve(request.user)
                    processed += 1
                elif action == "reject":
                    mobile_import.reject(request.user, "Bulk rejected")
                    processed += 1
            except Exception as e:
                logger = __import__('json').logger or print
                logger.error(f"Failed to process import {mobile_import.id}: {e}")
                continue

        messages.success(request, f"{processed} imports processed successfully.")

    return redirect("locations:mobile_import_list")


@login_required
@staff_required
def bulk_census_import_list(request):
    """List all bulk census import requests with comprehensive statistics"""
    if request.user.role in ['ADMIN', 'SUPERADMIN']:
        imports = BulkCensusImport.objects.all()
    else:
        imports = BulkCensusImport.objects.filter(uploaded_by=request.user)

    # Filter by status
    status_filter = request.GET.get("status", "")
    if status_filter:
        imports = imports.filter(status=status_filter)

    # Filter by file type
    type_filter = request.GET.get("type", "")
    if type_filter:
        imports = imports.filter(file_type=type_filter)

    # Order by most recent first
    imports = imports.select_related('uploaded_by', 'processed_by').order_by('-uploaded_at')

    from django.core.paginator import Paginator
    paginator = Paginator(imports, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Calculate statistics
    stats = {
        'total': imports.count(),
        'pending': imports.filter(status='pending').count(),
        'validating': imports.filter(status='validating').count(),
        'ready': imports.filter(status='ready').count(),
        'importing': imports.filter(status='importing').count(),
        'completed': imports.filter(status='completed').count(),
        'failed': imports.filter(status='failed').count(),
    }

    context = {
        "page_obj": page_obj,
        "status_filter": status_filter,
        "type_filter": type_filter,
        "status_choices": BulkCensusImport.IMPORT_STATUS,
        "type_choices": BulkCensusImport.IMPORT_TYPES,
        "stats": stats,
        "is_admin": request.user.role in ['ADMIN', 'SUPERADMIN'],
    }

    return render(request, "locations/bulk_import_list.html", context)


@login_required
@staff_required
def bulk_census_import_upload(request):
    """Upload and validate bulk census data file"""
    if request.method == "POST" and request.FILES.get('file'):
        file = request.FILES['file']

        # Validate file type
        if not file.name.endswith(('.xlsx', '.xls', '.csv')):
            messages.error(request, "Please upload an Excel (.xlsx, .xls) or CSV file.")
            return redirect("locations:bulk_census_import_list")

        # Validate file size (max 50MB for bulk imports)
        if file.size > 50 * 1024 * 1024:
            messages.error(request, "File size must be less than 50MB.")
            return redirect("locations:bulk_census_import_list")

        try:
            # Create bulk import record
            bulk_import = BulkCensusImport.objects.create(
                file=file,
                file_type="excel" if file.name.endswith(('.xlsx', '.xls')) else "csv",
                original_filename=file.name,
                uploaded_by=request.user,
                total_rows=0,  # Will be updated during validation
            )

            # Start validation in background (for now, do it synchronously)
            validate_bulk_import(bulk_import.id)

            messages.success(
                request,
                f"File '{file.name}' uploaded successfully. Validation in progress..."
            )

        except Exception as e:
            messages.error(request, f"Upload failed: {str(e)}")

    return redirect("locations:bulk_census_import_list")


@login_required
@staff_required
def bulk_census_import_review(request, import_id):
    """Review bulk import with validation results and preview"""
    bulk_import = get_object_or_404(BulkCensusImport, id=import_id)

    # Check if user can review this import
    if not (bulk_import.uploaded_by == request.user or request.user.role in ['ADMIN', 'SUPERADMIN']):
        messages.error(request, "You don't have permission to review this import.")
        return redirect("locations:bulk_census_import_list")

    # If import is not ready, redirect to list
    if bulk_import.status not in ['ready', 'validation_failed', 'completed', 'failed']:
        messages.info(request, "Import is still being processed. Please wait.")
        return redirect("locations:bulk_census_import_list")

    # Get validation errors for display
    validation_errors = bulk_import.validation_errors if bulk_import.status == 'validation_failed' else []

    context = {
        "bulk_import": bulk_import,
        "validation_errors": validation_errors,
        "can_process": bulk_import.can_process(request.user),
        "is_admin": request.user.role in ['ADMIN', 'SUPERADMIN'],
    }

    return render(request, "locations/bulk_import_review.html", context)


@login_required
@staff_required
def bulk_census_import_process(request, import_id):
    """Process and import validated bulk census data"""
    bulk_import = get_object_or_404(BulkCensusImport, id=import_id)

    # Check permissions
    if not bulk_import.can_process(request.user):
        messages.error(request, "You don't have permission to process this import.")
        return redirect("locations:bulk_census_import_list")

    if request.method == "POST":
        try:
            # Process the import
            imported_count = process_bulk_import(bulk_import.id, request.user)

            # Log activity
            log_census_activity(
                census=None,  # Bulk import doesn't have a single census
                activity_type="bulk_import",
                description=f"Bulk imported {imported_count} census records from {bulk_import.original_filename}",
                user=request.user,
                request=request
            )

            messages.success(
                request,
                f"Successfully imported {imported_count} census records from {bulk_import.original_filename}"
            )

        except Exception as e:
            bulk_import.mark_as_failed(e)
            messages.error(request, f"Import failed: {str(e)}")

    return redirect("locations:bulk_census_import_list")


@login_required
@staff_required
def bulk_census_import_cancel(request, import_id):
    """Cancel a bulk import"""
    bulk_import = get_object_or_404(BulkCensusImport, id=import_id)

    # Check permissions
    if not (bulk_import.uploaded_by == request.user or request.user.role in ['ADMIN', 'SUPERADMIN']):
        messages.error(request, "You don't have permission to cancel this import.")
        return redirect("locations:bulk_census_import_list")

    if request.method == "POST":
        bulk_import.status = "cancelled"
        bulk_import.save()

        messages.success(request, f"Bulk import '{bulk_import.original_filename}' cancelled.")

    return redirect("locations:bulk_census_import_list")


# Background processing functions (simplified for now)
def validate_bulk_import(import_id):
    """Validate bulk import file and prepare preview data"""
    try:
        bulk_import = BulkCensusImport.objects.get(id=import_id)
        bulk_import.mark_as_validating()

        # Read and validate file
        if bulk_import.file_type == "excel":
            data = read_excel_file(bulk_import.file.path)
        else:
            data = read_csv_file(bulk_import.file.path)

        # Validate data structure and content
        validation_result = validate_import_data(data, bulk_import)

        if validation_result["valid"]:
            # Prepare preview data
            preview_data = prepare_import_preview(data, bulk_import)
            bulk_import.mark_as_ready(preview_data)
        else:
            bulk_import.mark_validation_failed(validation_result["errors"])

    except Exception as e:
        bulk_import.mark_as_failed(e)


def process_bulk_import(import_id, user):
    """Process and import bulk census data"""
    try:
        bulk_import = BulkCensusImport.objects.get(id=import_id)
        bulk_import.mark_as_importing()

        # Read file data
        if bulk_import.file_type == "excel":
            data = read_excel_file(bulk_import.file.path)
        else:
            data = read_csv_file(bulk_import.file.path)

        # Import the data
        imported_count = import_census_data(data, bulk_import, user)

        bulk_import.mark_as_completed(imported_count)
        return imported_count

    except Exception as e:
        bulk_import.mark_as_failed(e)
        raise e


def read_excel_file(file_path):
    """Read Excel file and return structured data"""
    workbook = load_workbook(file_path, read_only=True)
    sheet = workbook.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]

    data = []
    for row in sheet.iter_rows(min_row=2):
        row_data = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = cell.value
        data.append(row_data)

    return {"headers": headers, "data": data}


def read_csv_file(file_path):
    """Read CSV file and return structured data"""
    import csv

    data = []
    with open(file_path, 'r', encoding='utf-8') as file:
        reader = csv.DictReader(file)
        for row in reader:
            data.append(row)

    return {"headers": list(data[0].keys()) if data else [], "data": data}


def validate_import_data(data, bulk_import):
    """Validate import data structure and content"""
    errors = []
    valid_rows = 0

    required_columns = ['site', 'species', 'count', 'date']
    headers = data.get('headers', [])

    # Check required columns
    missing_columns = [col for col in required_columns if col not in headers]
    if missing_columns:
        errors.append(f"Missing required columns: {', '.join(missing_columns)}")

    if errors:
        return {"valid": False, "errors": errors}

    # Validate each row
    for i, row in enumerate(data.get('data', [])):
        row_errors = []

        # Validate site
        site_name = str(row.get('site', '')).strip()
        if not site_name:
            row_errors.append("Site name is required")
        else:
            # Check if site exists or suggest creation
            try:
                Site.objects.get(name__iexact=site_name)
            except Site.DoesNotExist:
                row_errors.append(f"Site '{site_name}' does not exist in database")

        # Validate species
        species_name = str(row.get('species', '')).strip()
        if not species_name:
            row_errors.append("Species name is required")

        # Validate count
        try:
            count = int(row.get('count', 0))
            if count < 0:
                row_errors.append("Count must be non-negative")
        except (ValueError, TypeError):
            row_errors.append("Count must be a valid number")

        # Validate date
        date_str = str(row.get('date', '')).strip()
        if not date_str:
            row_errors.append("Date is required")
        else:
            # Basic date validation
            try:
                from datetime import datetime
                datetime.strptime(date_str, '%Y-%m-%d')
            except ValueError:
                row_errors.append("Date must be in YYYY-MM-DD format")

        if row_errors:
            errors.append({
                "row": i + 2,  # +2 because row 1 is headers, indexing starts at 0
                "errors": row_errors,
                "data": row
            })
        else:
            valid_rows += 1

    # Update bulk import with validation results
    bulk_import.total_rows = len(data.get('data', []))
    bulk_import.valid_rows = valid_rows
    bulk_import.invalid_rows = len(errors)
    bulk_import.save()

    return {"valid": len(errors) == 0, "errors": errors}


def prepare_import_preview(data, bulk_import):
    """Prepare preview data for the import"""
    preview_data = []

    for row in data.get('data', [])[:10]:  # Show first 10 rows
        preview_data.append({
            "site": row.get('site', ''),
            "species": row.get('species', ''),
            "count": row.get('count', ''),
            "date": row.get('date', ''),
            "notes": row.get('notes', '') if 'notes' in data.get('headers', []) else '',
        })

    return preview_data


def import_census_data(data, bulk_import, user):
    """Import validated census data"""
    imported_count = 0

    for row in data.get('data', []):
        try:
            # Get or create site
            site_name = str(row.get('site', '')).strip()
            site, created = Site.objects.get_or_create(
                name=site_name,
                defaults={
                    'site_type': 'other',
                    'description': f'Created from bulk import: {bulk_import.original_filename}',
                    'created_by': user
                }
            )

            # Get or create species
            species_name = str(row.get('species', '')).strip()
            species, created = Species.objects.get_or_create(
                name=species_name,
                defaults={
                    'scientific_name': species_name,  # Use common name as scientific for now
                    'iucn_status': 'LC'  # Default to Least Concern
                }
            )

            # Parse date
            date_str = str(row.get('date', '')).strip()
            from datetime import datetime
            observation_date = datetime.strptime(date_str, '%Y-%m-%d').date()

            # Check for existing census observation
            existing_census = CensusObservation.objects.filter(
                site=site,
                observation_date=observation_date
            ).first()

            if existing_census:
                # Add to existing census
                SpeciesObservation.objects.create(
                    census=existing_census,
                    species=species,
                    species_name=species_name,
                    count=int(row.get('count', 0)),
                    behavior_notes=str(row.get('notes', '')) if 'notes' in data.get('headers', []) else ''
                )
            else:
                # Create new census observation
                census = CensusObservation.objects.create(
                    site=site,
                    observation_date=observation_date,
                    observer=user,
                    weather_conditions="Imported from historical data",
                    notes=f"Bulk imported from {bulk_import.original_filename}"
                )

                SpeciesObservation.objects.create(
                    census=census,
                    species=species,
                    species_name=species_name,
                    count=int(row.get('count', 0)),
                    behavior_notes=str(row.get('notes', '')) if 'notes' in data.get('headers', []) else ''
                )

            imported_count += 1

        except Exception as e:
            # Log error but continue with other rows
            continue

    return imported_count
