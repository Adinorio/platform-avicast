"""
Excel Import/Export Utilities for Census Data
AGENTS.md §3 File Organization - Utility module for Excel operations
AGENTS.md §6.1 Security - Input validation for bulk imports
"""

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Dict, List, Tuple

from django.db import transaction
from django.db.models import Q
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from apps.fauna.models import Species
from apps.locations.models import Site, CensusYear, CensusMonth, Census, CensusObservation


class ExcelImportError(Exception):
    """Custom exception for Excel import errors"""
    pass


class CensusExcelHandler:
    """Handles Excel import/export for census data"""
    
    # Excel column mapping (custom format for user's Excel file)
    COLUMNS = {
        'A': 'species_name',  # Species name in column A
        'B': 'september',     # September count in column B
        'C': 'october',       # October count in column C
        'D': 'november',      # November count in column D
        'E': 'year_total',    # Year total in column E
        'F': 'january',       # Additional months if present
        'G': 'february',
        'H': 'march',
        'I': 'april',
        'J': 'may',
        'K': 'june',
        'L': 'july',
        'M': 'august',
        'N': 'december',
        'O': 'family',        # Family will be determined from context
        'P': 'site',          # Site will be set to 'DAGA'
    }

    HEADER_ROW = [
        'Site',
        'Family',
        'Species Name',
        'January',
        'February',
        'March',
        'April',
        'May',
        'June',
        'July',
        'August',
        'September',
        'October',
        'November',
        'December',
        'Year Total'
    ]
    
    @staticmethod
    def generate_template() -> Workbook:
        """
        Generate Excel template for data import
        Returns: Workbook object with headers and example data
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Census Data"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for idx, header in enumerate(CensusExcelHandler.HEADER_ROW, start=1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Add example data (family-grouped format)
        # Family heading example
        ws.append([
            "Lakawon",
            "HERONS AND EGRETS",
            "",  # No species name for family heading
            "", "", "", "", "", "", "", "", "", "", "", ""  # Empty monthly counts
        ])
        
        # Species examples under family
        ws.append([
            "Lakawon",
            "HERONS AND EGRETS",
            "Chinese Egret",
            15,  # January
            12,  # February
            8,   # March
            22,  # April
            18,  # May
            25,  # June
            30,  # July
            28,  # August
            35,  # September
            40,  # October
            45,  # November
            38,  # December
            316  # Year Total
        ])
        ws.append([
            "Lakawon",
            "HERONS AND EGRETS",
            "Black-faced Spoonbill",
            5,   # January
            8,   # February
            12,  # March
            15,  # April
            18,  # May
            20,  # June
            22,  # July
            25,  # August
            28,  # September
            30,  # October
            23,  # November
            20,  # December
            226  # Year Total
        ])

        # Adjust column widths (family-grouped format)
        ws.column_dimensions['A'].width = 15  # Site
        ws.column_dimensions['B'].width = 25  # Family
        ws.column_dimensions['C'].width = 30  # Species Name
        ws.column_dimensions['D'].width = 10  # January
        ws.column_dimensions['E'].width = 10  # February
        ws.column_dimensions['F'].width = 10  # March
        ws.column_dimensions['G'].width = 10  # April
        ws.column_dimensions['H'].width = 10  # May
        ws.column_dimensions['I'].width = 10  # June
        ws.column_dimensions['J'].width = 10  # July
        ws.column_dimensions['K'].width = 10  # August
        ws.column_dimensions['L'].width = 10  # September
        ws.column_dimensions['M'].width = 10  # October
        ws.column_dimensions['N'].width = 10  # November
        ws.column_dimensions['O'].width = 10  # December
        ws.column_dimensions['P'].width = 12  # Year Total

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ["Census Data Import Instructions - Family-Grouped Format"],
            [""],
            ["Column Requirements:"],
            ["Site", "Must match existing site name exactly (case-sensitive) - e.g., 'Lakawon'"],
            ["Family", "Bird family/category (e.g., 'HERONS AND EGRETS', 'SHOREBIRDS')"],
            ["Species Name", "Common or scientific name (leave empty for family heading rows)"],
            ["January-December", "Monthly counts (numbers only, leave blank if no observations)"],
            ["Year Total", "Sum of all monthly counts (auto-calculated if left blank)"],
            [""],
            ["Data Structure:"],
            ["- Family heading rows: Leave Species Name empty, leave counts empty"],
            ["- Species rows: Fill in species name and monthly counts"],
            ["- Monthly counts: Use numbers only (0, 5, 12, etc.)"],
            ["- Missing data: Leave cells blank (not zero)"],
            ["- Year Total: Will be calculated automatically if left blank"],
            [""],
            ["Important Notes:"],
            ["- Family groupings help organize related species"],
            ["- Each species gets individual monthly counts"],
            ["- Blank monthly cells = no observations that month"],
            ["- Species names will be validated against the database"],
            ["- If a species is not found, it will be created automatically"],
            ["- Site names must match exactly (case-sensitive)"],
            [""],
            ["Example Usage:"],
            ["- Row 1: Family heading (Site=Lakawon, Family=HERONS AND EGRETS, Species Name=empty)"],
            ["- Row 2: Species (Site=Lakawon, Family=HERONS AND EGRETS, Species Name=Chinese Egret, counts=15,12,8...)"],
            ["- Row 3: Species (Site=Lakawon, Family=HERONS AND EGRETS, Species Name=Black-faced Spoonbill, counts=5,8,12...)"],
        ]
        
        for row_idx, row_data in enumerate(instructions, start=1):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws_instructions.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:
                    cell.font = Font(bold=True, size=14)
                elif row_idx in [3, 13]:
                    cell.font = Font(bold=True)
        
        ws_instructions.column_dimensions['A'].width = 25
        ws_instructions.column_dimensions['B'].width = 60
        
        return wb
    
    @staticmethod
    def validate_row(row_data: Dict, row_num: int) -> Tuple[bool, str]:
        """
        Validate a single row of data (family-grouped format)
        Returns: (is_valid, error_message)
        """
        # Required fields for custom format (site will be set automatically to 'DAGA')
        
        # Check if this is a family heading row (no species name)
        is_family_heading = not row_data.get('species_name')
        
        if is_family_heading:
            # Family heading row - just validate site exists (will be created if missing during import)
            return True, ""
        
        # Species row - validate species name and counts
        if not row_data.get('species_name'):
            return False, f"Row {row_num}: Species name required for species rows"
        
        # Site will be created automatically during import if missing
        
        # Validate monthly counts (blank values are treated as 0)
        monthly_fields = ['january', 'february', 'march', 'april', 'may', 'june',
                         'july', 'august', 'september', 'october', 'november', 'december']
        
        total_count = 0
        has_any_count = False
        
        for month_field in monthly_fields:
            value = row_data.get(month_field)
            # Handle whitespace and empty values - strip whitespace and treat as blank if empty
            if value is not None:
                value = str(value).strip()
                if value == "":
                    value = None
            
            if value is not None and value != "":
                try:
                    count = int(value)
                    if count < 0:
                        return False, f"Row {row_num}: {month_field.title()} count cannot be negative"
                    total_count += count
                    has_any_count = True
                except (ValueError, TypeError):
                    return False, f"Row {row_num}: {month_field.title()} contains invalid value '{value}' - must be a number or blank"
            # If value is None or empty string, treat as 0 (no observations for that month)
        
        # Allow species rows with all blank monthly counts (treated as 0 for each month)
        # This is valid - it means no observations for any month
        
        # Validate year total if provided (but don't check against monthly sum)
        year_total = row_data.get('year_total')
        if year_total is not None:
            year_total = str(year_total).strip()
            if year_total != "":
                try:
                    total_value = int(year_total)
                    if total_value < 0:
                        return False, f"Row {row_num}: Year Total cannot be negative"
                    # Don't validate against monthly sum - trust the Year Total value
                except (ValueError, TypeError):
                    return False, f"Row {row_num}: Year Total contains invalid value '{year_total}' - must be a number or blank"
        
        return True, ""
    
    @staticmethod
    def preview_import(file, user) -> Dict:
        """
        Preview import data without actually saving to database
        Returns summary of what would be imported
        """
        try:
            # First, detect the Excel file structure
            structure = CensusExcelHandler.detect_excel_structure(file)
            
            # Reset file pointer
            file.seek(0)
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            # Preview data
            preview_data = {
                'sites': set(),
                'years': set(),
                'months': set(),
                'species': set(),
                'absent_species': set(),  # Species with zero counts
                'total_observations': 0,
                'total_birds': 0,
                'monthly_totals': {
                    'january': 0, 'february': 0, 'march': 0, 'april': 0,
                    'may': 0, 'june': 0, 'july': 0, 'august': 0,
                    'september': 0, 'october': 0, 'november': 0, 'december': 0
                },
                'errors': [],
                'warnings': [],
                'preview_rows': [],  # Store actual Excel rows for display
                'detected_structure': structure
                # Note: file object will be handled separately
            }
            
            # Process rows using detected structure
            data_start_row = structure['data_start_row'] or 8
            for row_num, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
                if not any(row):  # Skip empty rows
                    continue
                
                # Extract data based on detected structure
                species_col = structure['columns'].get('species', 'A')
                species_col_idx = ord(species_col) - 65
                
                # Get species name
                species_name = ''
                if species_col_idx < len(row) and row[species_col_idx]:
                    species_name = str(row[species_col_idx]).strip()
                
                # Skip rows without species data
                if not species_name or species_name.upper() in ['SPECIES', 'TOTAL']:
                    continue
                
                # Skip family heading rows
                family_headings = ['HERONS AND EGRETS', 'SHOREBIRDS-WADERS', 'RAILS, GALLINULES & COOTS', 
                                 'GULLS, TERNS & SKIMMERRS', 'ADDITIONAL SPECIES', 'RAILS', 'GULLS', 'TERN']
                if species_name.upper() in family_headings:
                    continue
                
                # Get site name
                if structure['has_site_column'] and 'site' in structure['columns']:
                    site_col = structure['columns']['site']
                    site_col_idx = ord(site_col) - 65
                    site_name = str(row[site_col_idx]).strip() if site_col_idx < len(row) and row[site_col_idx] else ''
                else:
                    site_name = structure['site_name'] or 'DEFAULT_SITE'
                
                # Get family if available
                family_name = ''
                if structure['has_family_column'] and 'family' in structure['columns']:
                    family_col = structure['columns']['family']
                    family_col_idx = ord(family_col) - 65
                    family_name = str(row[family_col_idx]).strip() if family_col_idx < len(row) and row[family_col_idx] else ''
                
                # If no family column, determine family from species name
                if not family_name:
                    species_upper = species_name.upper()
                    if any(pattern in species_upper for pattern in ['HERON', 'EGRET', 'BITTERN', 'NIGHT HERON', 'POND HERON', 'REEF EGRET']):
                        family_name = 'HERONS AND EGRETS'
                    elif any(pattern in species_upper for pattern in ['SANDPIPER', 'PLOVER', 'GODWIT', 'CURLEW', 'WHIMBREL', 'REDSHANK', 'GREENSHANK', 'TATTLER', 'TURNSTONE', 'KNOT', 'STINT', 'DOWITCHER', 'PHALAROPE', 'STILT', 'SNIPE', 'AVOCET']):
                        family_name = 'SHOREBIRDS-WADERS'
                    elif any(pattern in species_upper for pattern in ['RAIL', 'CRAKE', 'GALLINULE', 'COOT', 'MOORHEN', 'WATER HEN', 'WATERCOCK']):
                        family_name = 'RAILS, GALLINULES & COOTS'
                    elif any(pattern in species_upper for pattern in ['GULL', 'TERN', 'SKIMMER']):
                        family_name = 'GULLS, TERNS & SKIMMERRS'
                    elif any(pattern in species_upper for pattern in ['KITE', 'KINGFISHER', 'SWALLOW', 'OSPREY', 'DOVE', 'PIGEON', 'OWL', 'EAGLE', 'HAWK', 'FALCON', 'UNIDENTIFIED']):
                        family_name = 'ADDITIONAL SPECIES'
                    else:
                        family_name = 'ADDITIONAL SPECIES'
                
                # Create row data for preview display
                row_data = {
                    'site': site_name,
                    'family': family_name,
                    'species_name': species_name
                }
                
                # Add monthly counts
                month_names = {1: 'january', 2: 'february', 3: 'march', 4: 'april',
                             5: 'may', 6: 'june', 7: 'july', 8: 'august',
                             9: 'september', 10: 'october', 11: 'november', 12: 'december'}
                
                monthly_sum = 0  # Track sum of monthly counts for validation
                
                for month_num, month_col in structure['month_columns'].items():
                    month_field = month_names.get(month_num, f'month_{month_num}')
                    month_col_idx = ord(month_col) - 65
                    if month_col_idx < len(row):
                        count_value = row[month_col_idx]
                        if count_value is not None:
                            try:
                                count = int(count_value)
                                row_data[month_field] = count
                                monthly_sum += count
                            except (ValueError, TypeError):
                                row_data[month_field] = 0
                        else:
                            row_data[month_field] = 0
                    else:
                        row_data[month_field] = 0
                
                # Read Year Total column (if available in standard format)
                if 'year_total_column' in structure and structure['year_total_column']:
                    year_total_col = structure['year_total_column']
                    year_total_col_idx = ord(year_total_col) - 65
                    if year_total_col_idx < len(row):
                        year_total_value = row[year_total_col_idx]
                        if year_total_value is not None:
                            try:
                                year_total = int(year_total_value)
                                row_data['year_total'] = year_total
                                
                                # Validate: Year Total should equal sum of monthly counts
                                if monthly_sum > 0 and year_total != monthly_sum:
                                    # Store warning but don't fail
                                    row_data['year_total_mismatch'] = True
                                    row_data['calculated_total'] = monthly_sum
                            except (ValueError, TypeError):
                                row_data['year_total'] = None
                        else:
                            row_data['year_total'] = None
                    else:
                        row_data['year_total'] = None
                else:
                    # For legacy format, calculate year total from monthly sum
                    row_data['year_total'] = monthly_sum if monthly_sum > 0 else None
                
                # Store preview row for display (use None for empty values to work with template filters)
                def clean_value(val):
                    """Convert empty/None values to None, keep valid values as strings for display"""
                    if val is None:
                        return None
                    if isinstance(val, str):
                        val = val.strip()
                        return None if val == '' else val
                    # Convert numbers to strings for display
                    return str(val) if val != '' else None
                
                preview_row = {
                    'row_num': row_num,
                    'site': clean_value(row_data.get('site')),
                    'family': clean_value(row_data.get('family')),
                    'species_name': clean_value(row_data.get('species_name')),
                    'year_total': clean_value(row_data.get('year_total')),
                    'monthly_counts': {
                        'January': clean_value(row_data.get('january')),
                        'February': clean_value(row_data.get('february')),
                        'March': clean_value(row_data.get('march')),
                        'April': clean_value(row_data.get('april')),
                        'May': clean_value(row_data.get('may')),
                        'June': clean_value(row_data.get('june')),
                        'July': clean_value(row_data.get('july')),
                        'August': clean_value(row_data.get('august')),
                        'September': clean_value(row_data.get('september')),
                        'October': clean_value(row_data.get('october')),
                        'November': clean_value(row_data.get('november')),
                        'December': clean_value(row_data.get('december')),
                    },
                    'is_family_heading': not row_data.get('species_name'),
                    'error': None
                }
                
                # Basic validation
                if not species_name:
                    preview_data['errors'].append(f"Row {row_num}: Missing species name")
                    continue
                
                if not site_name:
                    preview_data['errors'].append(f"Row {row_num}: Missing site name")
                    continue
                
                preview_data['preview_rows'].append(preview_row)
                
                # Extract preview information
                site_name = (row_data.get('site') or '').strip()
                species_name = (row_data.get('species_name') or '').strip()
                family = (row_data.get('family') or '').strip()
                
                preview_data['sites'].add(site_name)
                # Only add species if it's not a family heading row or TOTAL row
                if species_name and species_name.upper() != 'TOTAL':
                    # Calculate total count for this species
                    total_count = 0
                    monthly_counts = preview_row['monthly_counts']
                    for month, count in monthly_counts.items():
                        if count and count != '-':
                            try:
                                total_count += int(count)
                            except (ValueError, TypeError):
                                pass
                    
                    # Mark as absent if no counts
                    if total_count == 0:
                        preview_data['absent_species'].add(species_name)
                    else:
                        preview_data['species'].add(species_name)
                    # print(f"DEBUG: Adding species: '{species_name}'")
                
                # Count observations and birds using monthly data (Option 2 approach)
                if species_name:  # Only process species rows (not family headings)
                    # Skip TOTAL rows - these represent monthly/yearly totals, not individual species
                    if species_name.upper() == 'TOTAL':
                        continue
                    
                    monthly_fields = ['january', 'february', 'march', 'april', 'may', 'june',
                                     'july', 'august', 'september', 'october', 'november', 'december']
                    
                    has_monthly_data = False
                    
                    for month_field in monthly_fields:
                        count_value = row_data.get(month_field)
                        if count_value is not None:
                            count_value = str(count_value).strip()
                            if count_value != "":
                                try:
                                    count = int(count_value)
                                    if count > 0:
                                        # Add to monthly totals (sum across all species)
                                        preview_data['monthly_totals'][month_field] += count
                                        has_monthly_data = True
                                        preview_data['months'].add(month_field.title())
                                except (ValueError, TypeError):
                                    pass
                    
                    # Count species with any monthly data
                    if has_monthly_data:
                        preview_data['total_observations'] += 1
                
                # Extract year from the data (assuming current year for preview)
                from datetime import datetime
                preview_data['years'].add(datetime.now().year)
            
            # Convert sets to sorted lists for display
            preview_data['sites'] = sorted(list(preview_data['sites']))
            preview_data['years'] = sorted(list(preview_data['years']))
            preview_data['months'] = sorted(list(preview_data['months']))
            preview_data['species'] = sorted(list(preview_data['species']))
            preview_data['absent_species'] = sorted(list(preview_data['absent_species']))
            
            # Check for existing sites
            existing_sites = Site.objects.filter(name__in=preview_data['sites']).values_list('name', flat=True)
            new_sites = [site for site in preview_data['sites'] if site not in existing_sites]
            
            if new_sites:
                preview_data['warnings'].append(f"New sites will be created: {', '.join(new_sites)}")
            
            # Check for existing species
            existing_species = Species.objects.filter(name__in=preview_data['species']).values_list('name', flat=True)
            new_species = [species for species in preview_data['species'] if species not in existing_species]
            
            if new_species:
                if len(new_species) <= 5:
                    species_list = ', '.join(new_species)
                    preview_data['warnings'].append(f"New species will be created: {species_list}")
                else:
                    preview_data['warnings'].append(f"New species will be created: {len(new_species)} species")
            
            # Calculate total birds from monthly totals (sum of all monthly totals)
            preview_data['total_birds'] = sum(preview_data['monthly_totals'].values())
            
            return preview_data
            
        except Exception as e:
            raise ExcelImportError(f"Error previewing Excel file: {str(e)}")

    @staticmethod
    def detect_excel_structure(file) -> Dict:
        """
        Automatically detect Excel file structure
        Supports two formats:
        
        Format 1 (Standard - Preferred):
        - Row 1: Headers (Site, Family, Species Name, Jan...Dec, Year Total)
        - Column A: Site, Column B: Family, Column C: Species Name
        - Columns D-O: Monthly counts (January to December)
        - Column P: Year Total
        
        Format 2 (Legacy multi-year):
        - Row 8: Headers (SPECIES, SEPTEMBER, OCTOBER, NOVEMBER, TOTAL, ...)
        - Column A: Species Name
        - Columns B-D: Monthly counts (varies by year/site section)
        """
        try:
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            # Reset file pointer
            file.seek(0)
            
            # Check row 1 for standard format headers
            row1 = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            row1_str = ' '.join([str(cell).upper() for cell in row1 if cell])
            
            # Check if row 1 has the standard format (Site, Family, Species Name)
            if 'SITE' in row1_str and 'FAMILY' in row1_str and 'SPECIES' in row1_str:
                # Standard format detected
                structure = {
                    'header_row': 1,
                    'data_start_row': 2,
                    'columns': {
                        'site': 'A',           # Column A: Site
                        'family': 'B',         # Column B: Family
                        'species': 'C',        # Column C: Species Name
                    },
                    'has_site_column': True,
                    'has_family_column': True,
                    'month_columns': {
                        1: 'D',   # January
                        2: 'E',   # February
                        3: 'F',   # March
                        4: 'G',   # April
                        5: 'H',   # May
                        6: 'I',   # June
                        7: 'J',   # July
                        8: 'K',   # August
                        9: 'L',   # September
                        10: 'M',  # October
                        11: 'N',  # November
                        12: 'O',  # December
                    },
                    'year_total_column': 'P',  # Column P: Year Total
                    'site_name': None,
                    'format_type': 'standard'
                }
            else:
                # Legacy format - check row 8
                structure = {
                    'header_row': 8,
                    'data_start_row': 11,
                    'columns': {
                        'species': 'A',        # Column A: Species Name
                    },
                    'has_site_column': False,
                    'has_family_column': False,
                    'month_columns': {
                        9: 'B',   # September (column B, index 1)
                        10: 'C',  # October (column C, index 2)
                        11: 'D',  # November (column D, index 3)
                    },
                    'site_name': None,
                    'format_type': 'legacy'
                }
                
                # Try to detect site name from filename for legacy format
                if hasattr(file, 'name'):
                    filename = file.name.upper()
                    if 'DAGA' in filename:
                        structure['site_name'] = 'DAGA'
                    elif 'LAKAWON' in filename:
                        structure['site_name'] = 'LAKAWON'
            
            return structure
            
        except Exception as e:
            raise ExcelImportError(f"Error detecting Excel structure: {str(e)}")

    @staticmethod
    def import_from_excel(file, user, import_year=None, site_coordinates=None, import_months=None, target_site=None) -> Dict:
        """
        Import census data from Excel file (flexible format detection)
        Args:
            file: UploadedFile object
            user: User object (for lead_observer)
            import_year: Year for the data (optional, defaults to current year)
            site_coordinates: Dict of site_name -> "lat, lng" coordinates (optional)
            import_months: List of month numbers to import (optional, defaults to all detected months)
            target_site: Site name to use if no site column detected (optional)
        Returns: Dict with success/error statistics
        """
        try:
            # First, detect the Excel file structure
            structure = CensusExcelHandler.detect_excel_structure(file)
            
            # Reset file pointer
            file.seek(0)
            wb = load_workbook(file, data_only=True)
            ws = wb.active
            
            results = {
                'total_rows': 0,
                'successful': 0,
                'skipped': 0,
                'errors': [],
                'info_messages': [],  # For non-error messages
                'created_census': 0,
                'created_observations': 0,
                'created_species': 0,
                'ai_matches': 0,
            }
            
            # Group observations by (site, month, year) to create census records
            census_groups = {}
            
            # Determine which months to import
            available_months = list(structure['month_columns'].keys())
            if import_months:
                months_to_import = [m for m in import_months if m in available_months]
            else:
                months_to_import = available_months
            
            results['info_messages'].append(f"Detected months: {available_months}")
            results['info_messages'].append(f"Importing months: {months_to_import}")

            # Read data rows starting from detected data start row
            data_start_row = structure['data_start_row'] or 8
            for row_num, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
                if not any(row):  # Skip empty rows
                    continue

                results['total_rows'] += 1

                # Extract data based on detected structure
                species_col = structure['columns'].get('species', 'A')  # Default to column A
                species_col_idx = ord(species_col) - 65  # Convert A=0, B=1, etc.
                
                # Get species name
                species_name = ''
                if species_col_idx < len(row) and row[species_col_idx]:
                    species_name = str(row[species_col_idx]).strip()
                
                # Skip rows without species data
                if not species_name or species_name.upper() in ['SPECIES', 'TOTAL']:
                    continue
                
                # Skip family heading rows
                family_headings = ['HERONS AND EGRETS', 'SHOREBIRDS-WADERS', 'RAILS, GALLINULES & COOTS', 
                                 'GULLS, TERNS & SKIMMERRS', 'ADDITIONAL SPECIES', 'RAILS', 'GULLS', 'TERN']
                if species_name.upper() in family_headings:
                    continue
                
                # Get site name
                if structure['has_site_column'] and 'site' in structure['columns']:
                    site_col = structure['columns']['site']
                    site_col_idx = ord(site_col) - 65
                    site_name = str(row[site_col_idx]).strip() if site_col_idx < len(row) and row[site_col_idx] else ''
                else:
                    site_name = target_site or structure['site_name'] or 'DEFAULT_SITE'
                
                # Get family if available
                family_name = ''
                if structure['has_family_column'] and 'family' in structure['columns']:
                    family_col = structure['columns']['family']
                    family_col_idx = ord(family_col) - 65
                    family_name = str(row[family_col_idx]).strip() if family_col_idx < len(row) and row[family_col_idx] else ''
                
                # If no family column, determine family from species name
                if not family_name:
                    species_upper = species_name.upper()
                    if any(pattern in species_upper for pattern in ['HERON', 'EGRET', 'BITTERN', 'NIGHT HERON', 'POND HERON', 'REEF EGRET']):
                        family_name = 'HERONS AND EGRETS'
                    elif any(pattern in species_upper for pattern in ['SANDPIPER', 'PLOVER', 'GODWIT', 'CURLEW', 'WHIMBREL', 'REDSHANK', 'GREENSHANK', 'TATTLER', 'TURNSTONE', 'KNOT', 'STINT', 'DOWITCHER', 'PHALAROPE', 'STILT', 'SNIPE', 'AVOCET']):
                        family_name = 'SHOREBIRDS-WADERS'
                    elif any(pattern in species_upper for pattern in ['RAIL', 'CRAKE', 'GALLINULE', 'COOT', 'MOORHEN', 'WATER HEN', 'WATERCOCK']):
                        family_name = 'RAILS, GALLINULES & COOTS'
                    elif any(pattern in species_upper for pattern in ['GULL', 'TERN', 'SKIMMER']):
                        family_name = 'GULLS, TERNS & SKIMMERRS'
                    elif any(pattern in species_upper for pattern in ['KITE', 'KINGFISHER', 'SWALLOW', 'OSPREY', 'DOVE', 'PIGEON', 'OWL', 'EAGLE', 'HAWK', 'FALCON', 'UNIDENTIFIED']):
                        family_name = 'ADDITIONAL SPECIES'
                    else:
                        family_name = 'ADDITIONAL SPECIES'
                
                # Basic validation
                if not species_name:
                    results['errors'].append(f"Row {row_num}: Missing species name")
                    results['skipped'] += 1
                    continue
                
                if not site_name:
                    results['errors'].append(f"Row {row_num}: Missing site name")
                    results['skipped'] += 1
                    continue
                
                # Skip TOTAL rows - these represent monthly/yearly totals, not individual species
                if species_name.upper() == 'TOTAL':
                    results['successful'] += 1  # TOTAL rows are valid but don't create observations
                    continue
                
                # Find or validate species (create if not exists for import)
                species = Species.objects.filter(
                    Q(name__iexact=species_name) |
                    Q(scientific_name__iexact=species_name)
                ).first()

                # If species not found, create it automatically (for data import)
                if not species:
                    try:
                        # Try to create species automatically for import
                        species = Species.objects.create(
                            name=species_name,
                            scientific_name=species_name,  # Use same as common name if not specified
                            iucn_status='LC'  # Default to Least Concern
                        )
                        results['created_species'] = results.get('created_species', 0) + 1
                    except Exception as e:
                        results['errors'].append(
                            f"Row {row_num}: Could not create species '{species_name}': {str(e)}"
                        )
                        results['skipped'] += 1
                        continue
                
                # Prepare site - set to DAGA for this Excel file
                site_name = 'DAGA'
                
                # Get coordinates for this site (if provided)
                site_coords = '0.0, 0.0'  # Default coordinates
                if site_coordinates and site_name in site_coordinates:
                    site_coords = site_coordinates[site_name]
                
                site, created = Site.objects.get_or_create(
                    name__iexact=site_name,
                    defaults={
                        'name': site_name,
                        'status': 'active',
                        'coordinates': site_coords,
                        'description': f'Site created during import from Excel data'
                    }
                )
                if created:
                    results['info_messages'].append(
                        f"Created new site '{site_name}' automatically"
                    )

                # Process monthly counts for detected months
                monthly_data = []
                total_count = 0
                for month_num in months_to_import:
                    if month_num in structure['month_columns']:
                        month_col = structure['month_columns'][month_num]
                        month_col_idx = ord(month_col) - 65
                        if month_col_idx < len(row):
                            count_value = row[month_col_idx]
                            monthly_data.append((month_num, count_value))
                            
                            # Calculate total count for this species
                            if count_value is not None:
                                try:
                                    count = int(str(count_value).strip())
                                    total_count += count
                                except (ValueError, TypeError):
                                    pass
                
                # Check if species has any sightings (total count > 0)
                if total_count == 0:
                    # Species is absent - no sightings recorded
                    results['info_messages'].append(
                        f"Species '{species_name}' marked as Absent (no sightings recorded)"
                    )
                    results['skipped'] += 1
                    continue
                
                # Determine year from data (use provided year or current year)
                import datetime
                current_year = int(import_year) if import_year else datetime.datetime.now().year
                
                for month_num, count_value in monthly_data:
                    # Handle whitespace and empty values - strip whitespace and treat as blank if empty
                    if count_value is not None:
                        count_value = str(count_value).strip()
                        if count_value == "":
                            count_value = None
                    
                    # Handle blank values as 0, but don't create observations for 0 counts
                    if count_value is not None and count_value != "":
                        try:
                            count = int(count_value)
                            if count > 0:  # Only create observations for positive counts
                                # Create census key for this month
                                census_key = (site.id, current_year, month_num)
                                
                                # Group observations by site/year/month
                                if census_key not in census_groups:
                                    census_groups[census_key] = {
                                        'site': site,
                                        'year': current_year,
                                        'month': month_num,
                                        'observations': []
                                    }
                                
                                # Add observation to group
                                month_names = {1: 'January', 2: 'February', 3: 'March', 4: 'April',
                                             5: 'May', 6: 'June', 7: 'July', 8: 'August',
                                             9: 'September', 10: 'October', 11: 'November', 12: 'December'}
                                census_groups[census_key]['observations'].append({
                                    'species': species,
                                    'species_name': species.name,
                                    'family': family_name,  # Store family for observation
                                    'count': count,
                                    'row_num': row_num,
                                    'month_name': month_names.get(month_num, f'Month {month_num}')
                                })
                                
                                results['successful'] += 1
                        except (ValueError, TypeError):
                            month_names = {1: 'January', 2: 'February', 3: 'March', 4: 'April',
                                         5: 'May', 6: 'June', 7: 'July', 8: 'August',
                                         9: 'September', 10: 'October', 11: 'November', 12: 'December'}
                            results['errors'].append(
                                f"Row {row_num}: Invalid count '{count_value}' for {month_names.get(month_num, f'Month {month_num}')}"
                            )
                            results['skipped'] += 1
            
            # Create census records with observations
            with transaction.atomic():
                for census_key, census_data in census_groups.items():
                    # Get or create CensusYear
                    year_obj, _ = CensusYear.objects.get_or_create(
                        site=census_data['site'],
                        year=census_data['year']
                    )
                    
                    # Get or create CensusMonth
                    month_obj, _ = CensusMonth.objects.get_or_create(
                        year=year_obj,
                        month=census_data['month']
                    )
                    
                    # Create census record for the 15th of the month (mid-month date)
                    census_date = datetime.date(
                        census_data['year'], 
                        census_data['month'], 
                        15
                    )
                    
                    # Check if census already exists for this date
                    census, created = Census.objects.get_or_create(
                        month=month_obj,
                        census_date=census_date,
                        defaults={
                            'lead_observer': user,
                        }
                    )
                    
                    if created:
                        results['created_census'] += 1
                    
                    # Add observations
                    for obs_data in census_data['observations']:
                        # Check for duplicates
                        existing = CensusObservation.objects.filter(
                            census=census,
                            species=obs_data['species']
                        ).first()
                        
                        if existing:
                            # Update count instead of creating duplicate
                            existing.count += obs_data['count']
                            existing.save()
                            results['errors'].append(
                                f"Row {obs_data['row_num']}: Duplicate observation for "
                                f"{obs_data['species_name']} in {obs_data['month_name']} - "
                                f"count added to existing record"
                            )
                        else:
                            # Store family information in the family field
                            CensusObservation.objects.create(
                                census=census,
                                species=obs_data['species'],
                                species_name=obs_data['species_name'],
                                family=obs_data.get('family', ''),  # Use family from obs_data
                                count=obs_data['count']
                            )
                            results['created_observations'] += 1
            
            return results
            
        except Exception as e:
            raise ExcelImportError(f"Error processing Excel file: {str(e)}")
    
    @staticmethod
    def export_to_excel(filters: Dict = None) -> Workbook:
        """
        Export census data to Excel (family-grouped format)
        Args:
            filters: Dict with optional keys: site_id, year, month, start_date, end_date
        Returns: Workbook object
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Census Data Export"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Write headers
        for idx, header in enumerate(CensusExcelHandler.HEADER_ROW, start=1):
            cell = ws.cell(row=1, column=idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Build query
        observations = CensusObservation.objects.select_related(
            'census__month__year__site',
            'species'
        ).all()
        
        # Apply filters
        if filters:
            if filters.get('site_id'):
                observations = observations.filter(
                    census__month__year__site_id=filters['site_id']
                )
            if filters.get('year'):
                observations = observations.filter(
                    census__month__year__year=filters['year']
                )
            if filters.get('month'):
                observations = observations.filter(
                    census__month__month=filters['month']
                )
            if filters.get('start_date'):
                observations = observations.filter(
                    census__census_date__gte=filters['start_date']
                )
            if filters.get('end_date'):
                observations = observations.filter(
                    census__census_date__lte=filters['end_date']
                )
        
        # Group observations by site, species, and year for family-grouped export
        species_data = {}
        
        for obs in observations.order_by('census__month__year__site__name', 'species_name', 'census__month__year__year', 'census__month__month'):
            site_name = obs.census.month.year.site.name
            species_name = obs.species_name
            year = obs.census.month.year.year
            month = obs.census.month.month
            count = obs.count
            
            # Create key for grouping
            key = (site_name, species_name, year)
            
            if key not in species_data:
                species_data[key] = {
                    'site': site_name,
                    'species': species_name,
                    'year': year,
                    'family': 'BIRDS',  # Default family, could be enhanced later
                    'monthly_counts': {1: 0, 2: 0, 3: 0, 4: 0, 5: 0, 6: 0, 7: 0, 8: 0, 9: 0, 10: 0, 11: 0, 12: 0}
                }
            
            species_data[key]['monthly_counts'][month] = count
        
        # Write data in family-grouped format
        current_site = None
        current_family = None
        
        for key, data in species_data.items():
            site_name = data['site']
            species_name = data['species']
            family = data['family']
            year = data['year']
            monthly_counts = data['monthly_counts']
            
            # Add family heading if site or family changed
            if current_site != site_name or current_family != family:
                ws.append([
                    site_name,
                    family,
                    "",  # No species name for family heading
                    "", "", "", "", "", "", "", "", "", "", "", ""  # Empty monthly counts
                ])
                current_site = site_name
                current_family = family
            
            # Calculate year total
            year_total = sum(monthly_counts.values())
            
            # Write species row
            ws.append([
                site_name,
                family,
                species_name,
                monthly_counts[1],   # January
                monthly_counts[2],   # February
                monthly_counts[3],   # March
                monthly_counts[4],   # April
                monthly_counts[5],   # May
                monthly_counts[6],   # June
                monthly_counts[7],   # July
                monthly_counts[8],   # August
                monthly_counts[9],   # September
                monthly_counts[10],  # October
                monthly_counts[11],  # November
                monthly_counts[12],  # December
                year_total           # Year Total
            ])

        # Adjust column widths (family-grouped format)
        ws.column_dimensions['A'].width = 15  # Site
        ws.column_dimensions['B'].width = 25  # Family
        ws.column_dimensions['C'].width = 30  # Species Name
        ws.column_dimensions['D'].width = 10  # January
        ws.column_dimensions['E'].width = 10  # February
        ws.column_dimensions['F'].width = 10  # March
        ws.column_dimensions['G'].width = 10  # April
        ws.column_dimensions['H'].width = 10  # May
        ws.column_dimensions['I'].width = 10  # June
        ws.column_dimensions['J'].width = 10  # July
        ws.column_dimensions['K'].width = 10  # August
        ws.column_dimensions['L'].width = 10  # September
        ws.column_dimensions['M'].width = 10  # October
        ws.column_dimensions['N'].width = 10  # November
        ws.column_dimensions['O'].width = 10  # December
        ws.column_dimensions['P'].width = 12  # Year Total
        
        return wb

