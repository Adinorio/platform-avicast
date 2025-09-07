import csv
import io
import json
from functools import wraps

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Sum
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook, load_workbook

from .forms import (
    CensusImportForm,
    CensusObservationForm,
    SiteForm,
    SpeciesObservationFormSet,
)
from .models import CensusObservation, MobileDataImport, Site, SpeciesObservation


def role_required(allowed_roles):
    """
    Decorator to check if user has required role
    """

    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect("login")

            if not hasattr(request.user, "role"):
                return HttpResponse("User role not defined", status=403)

            if request.user.role not in allowed_roles:
                return HttpResponse("Access denied. Insufficient permissions.", status=403)

            return view_func(request, *args, **kwargs)

        return _wrapped_view

    return decorator


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def site_list(request):
    """Display list of all sites with search and filtering"""
    sites = Site.objects.all()

    # Search functionality
    search_query = request.GET.get("search", "")
    if search_query:
        sites = sites.filter(
            Q(name__icontains=search_query)
            | Q(description__icontains=search_query)
            | Q(site_type__icontains=search_query)
        )

    # Filter by site type
    site_type = request.GET.get("site_type", "")
    if site_type:
        sites = sites.filter(site_type=site_type)

    # Filter by status
    status = request.GET.get("status", "")
    if status:
        sites = sites.filter(status=status)

    # Pagination
    paginator = Paginator(sites, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "search_query": search_query,
        "site_type": site_type,
        "status": status,
        "site_types": Site.SITE_TYPES,
        "site_statuses": Site.SITE_STATUS,
    }
    return render(request, "locations/site_list.html", context)


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def site_detail(request, site_id):
    """Display detailed site information with integrated census data"""
    site = get_object_or_404(Site, id=site_id)

    # Get census summary organized by year and month
    census_summary = site.get_census_summary()

    # Get recent census observations
    recent_census = CensusObservation.objects.filter(site=site).order_by("-observation_date")[:10]

    # Get species diversity stats
    total_species_observed = (
        SpeciesObservation.objects.filter(census__site=site)
        .values("species_name")
        .distinct()
        .count()
    )

    total_birds_observed = (
        SpeciesObservation.objects.filter(census__site=site).aggregate(total=Sum("count"))["total"]
        or 0
    )

    context = {
        "site": site,
        "census_summary": census_summary,
        "recent_census": recent_census,
        "total_species_observed": total_species_observed,
        "total_birds_observed": total_birds_observed,
    }
    return render(request, "locations/site_detail.html", context)


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def site_create(request):
    """Create a new site"""
    if request.method == "POST":
        form = SiteForm(request.POST)
        if form.is_valid():
            site = form.save(commit=False)
            site.created_by = request.user
            site.save()
            messages.success(request, f'Site "{site.name}" created successfully!')
            return redirect("locations:site_detail", site_id=site.id)
    else:
        form = SiteForm()

    context = {"form": form, "action": "Create"}
    return render(request, "locations/site_form.html", context)


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def site_update(request, site_id):
    """Update an existing site"""
    site = get_object_or_404(Site, id=site_id)

    if request.method == "POST":
        form = SiteForm(request.POST, instance=site)
        if form.is_valid():
            form.save()
            messages.success(request, f'Site "{site.name}" updated successfully!')
            return redirect("locations:site_detail", site_id=site.id)
    else:
        form = SiteForm(instance=site)

    context = {"form": form, "site": site, "action": "Update"}
    return render(request, "locations/site_form.html", context)


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def site_delete(request, site_id):
    """Delete a site"""
    site = get_object_or_404(Site, id=site_id)

    if request.method == "POST":
        site_name = site.name
        site.delete()
        messages.success(request, f'Site "{site_name}" deleted successfully!')
        return redirect("locations:site_list")

    context = {"site": site}
    return render(request, "locations/site_confirm_delete.html", context)


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def census_create(request, site_id):
    """Create a new census observation for a site"""
    site = get_object_or_404(Site, id=site_id)

    if request.method == "POST":
        census_form = CensusObservationForm(request.POST, site=site)
        species_formset = SpeciesObservationFormSet(request.POST, instance=census_form.instance)

        if census_form.is_valid() and species_formset.is_valid():
            with transaction.atomic():
                census = census_form.save(commit=False)
                census.site = site
                census.observer = request.user
                census.save()

                species_formset.instance = census
                species_formset.save()

                messages.success(
                    request, f"Census observation for {site.name} created successfully!"
                )
                return redirect("locations:site_detail", site_id=site.id)
    else:
        census_form = CensusObservationForm(site=site)
        species_formset = SpeciesObservationFormSet()

    context = {
        "site": site,
        "census_form": census_form,
        "species_formset": species_formset,
        "action": "Create",
    }
    return render(request, "locations/census_form.html", context)


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def census_update(request, census_id):
    """Update an existing census observation"""
    census = get_object_or_404(CensusObservation, id=census_id)

    if request.method == "POST":
        census_form = CensusObservationForm(request.POST, instance=census, site=census.site)
        species_formset = SpeciesObservationFormSet(request.POST, instance=census)

        if census_form.is_valid() and species_formset.is_valid():
            with transaction.atomic():
                census_form.save()
                species_formset.save()

                messages.success(request, "Census observation updated successfully!")
                return redirect("locations:site_detail", site_id=census.site.id)
    else:
        census_form = CensusObservationForm(instance=census, site=census.site)
        species_formset = SpeciesObservationFormSet(instance=census)

    context = {
        "census": census,
        "census_form": census_form,
        "species_formset": species_formset,
        "action": "Update",
    }
    return render(request, "locations/census_form.html", context)


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def census_delete(request, census_id):
    """Delete a census observation"""
    census = get_object_or_404(CensusObservation, id=census_id)
    site_id = census.site.id

    if request.method == "POST":
        census.delete()
        messages.success(request, "Census observation deleted successfully!")
        return redirect("locations:site_detail", site_id=site_id)

    context = {"census": census}
    return render(request, "locations/census_confirm_delete.html", context)


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def census_import(request, site_id):
    """Import census data from CSV/Excel file"""
    site = get_object_or_404(Site, id=site_id)

    if request.method == "POST":
        form = CensusImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Create census observation
                    census = CensusObservation.objects.create(
                        site=site,
                        observation_date=form.cleaned_data["observation_date"],
                        observer=request.user,
                        weather_conditions=form.cleaned_data["weather_conditions"],
                        notes=form.cleaned_data["notes"],
                    )

                    # Process the uploaded file
                    file = form.cleaned_data["file"]
                    if file.name.endswith(".csv"):
                        success_count = _process_csv_import(file, census)
                    else:
                        success_count = _process_excel_import(file, census)

                    messages.success(
                        request, f"Successfully imported {success_count} species observations!"
                    )
                    return redirect("locations:site_detail", site_id=site.id)

            except Exception as e:
                messages.error(request, f"Import failed: {str(e)}")
    else:
        form = CensusImportForm()

    context = {"form": form, "site": site}
    return render(request, "locations/census_import.html", context)


def _process_csv_import(file, census):
    """Process CSV file import"""
    decoded_file = file.read().decode("utf-8")
    csv_data = csv.DictReader(io.StringIO(decoded_file))

    success_count = 0
    for row in csv_data:
        try:
            SpeciesObservation.objects.create(
                census=census,
                species_name=row.get("species_name", "").strip(),
                count=int(row.get("count", 1)),
                behavior_notes=row.get("behavior_notes", "").strip(),
            )
            success_count += 1
        except (ValueError, KeyError):
            continue

    return success_count


def _process_excel_import(file, census):
    """Process Excel file import"""
    workbook = load_workbook(file)
    sheet = workbook.active

    success_count = 0
    headers = [cell.value for cell in sheet[1]]

    for row in sheet.iter_rows(min_row=2):
        try:
            row_data = {headers[i]: cell.value for i, cell in enumerate(row)}

            if row_data.get("species_name") and row_data.get("count"):
                SpeciesObservation.objects.create(
                    census=census,
                    species_name=str(row_data["species_name"]).strip(),
                    count=int(row_data["count"]),
                    behavior_notes=str(row_data.get("behavior_notes", "")).strip(),
                )
                success_count += 1
        except (ValueError, KeyError):
            continue

    return success_count


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def census_export_csv(request, site_id):
    """Export census data as CSV"""
    site = get_object_or_404(Site, id=site_id)

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{site.name}_census_data.csv"'

    writer = csv.writer(response)
    writer.writerow(["Date", "Species", "Count", "Behavior Notes", "Weather", "Observer"])

    census_observations = CensusObservation.objects.filter(site=site).order_by("-observation_date")

    for census in census_observations:
        for species_obs in census.species_observations.all():
            writer.writerow(
                [
                    census.observation_date,
                    species_obs.species_name,
                    species_obs.count,
                    species_obs.behavior_notes,
                    census.weather_conditions,
                    census.observer.username if census.observer else "",
                ]
            )

    return response


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def census_export_excel(request, site_id):
    """Export census data as Excel"""
    site = get_object_or_404(Site, id=site_id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"{site.name} Census Data"

    # Headers
    headers = ["Date", "Species", "Count", "Behavior Notes", "Weather", "Observer"]
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col, value=header)

    # Data
    row = 2
    census_observations = CensusObservation.objects.filter(site=site).order_by("-observation_date")

    for census in census_observations:
        for species_obs in census.species_observations.all():
            sheet.cell(row=row, column=1, value=census.observation_date)
            sheet.cell(row=row, column=2, value=species_obs.species_name)
            sheet.cell(row=row, column=3, value=species_obs.count)
            sheet.cell(row=row, column=4, value=species_obs.behavior_notes)
            sheet.cell(row=row, column=5, value=census.weather_conditions)
            sheet.cell(row=row, column=6, value=census.observer.username if census.observer else "")
            row += 1

    # Auto-adjust column widths
    for column in sheet.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (AttributeError, TypeError):
                pass

    # Mobile Data Import Views


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def mobile_import_list(request):
    """List all mobile data import requests"""
    if request.user.role == "ADMIN":
        imports = MobileDataImport.objects.all()
    else:
        imports = MobileDataImport.objects.filter(submitted_by=request.user)

    # Filter by status
    status_filter = request.GET.get("status", "")
    if status_filter:
        imports = imports.filter(status=status_filter)

    paginator = Paginator(imports, 20)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "page_obj": page_obj,
        "status_filter": status_filter,
        "status_choices": MobileDataImport.IMPORT_STATUS,
    }

    return render(request, "locations/mobile_import_list.html", context)


@login_required
@role_required(["ADMIN", "FIELD_WORKER"])
def submit_mobile_import(request):
    """Submit a new mobile data import request"""
    if request.method == "POST":
        try:
            import_data = json.loads(request.body)
            site_id = import_data.get("site_id")

            site = get_object_or_404(Site, id=site_id)

            # Create the import request
            mobile_import = MobileDataImport.objects.create(
                import_data=import_data, site=site, submitted_by=request.user, status="pending"
            )

            return JsonResponse(
                {
                    "success": True,
                    "message": "Import request submitted successfully",
                    "import_id": str(mobile_import.id),
                }
            )

        except Exception as e:
            return JsonResponse({"success": False, "message": f"Error submitting import: {str(e)}"})

    # GET request - show form
    sites = Site.objects.all()
    return render(request, "locations/submit_mobile_import.html", {"sites": sites})


@login_required
@role_required(["ADMIN"])
def review_mobile_import(request, import_id):
    """Review and approve/reject mobile data import requests"""
    mobile_import = get_object_or_404(MobileDataImport, id=import_id)

    if request.method == "POST":
        action = request.POST.get("action")
        notes = request.POST.get("review_notes", "")

        try:
            if action == "approve":
                mobile_import.approve(request.user)
                messages.success(request, "Import request approved successfully.")
            elif action == "reject":
                mobile_import.reject(request.user, notes)
                messages.success(request, "Import request rejected.")
            elif action == "process":
                census = mobile_import.process_import()
                messages.success(request, f"Import processed successfully. Census ID: {census.id}")

            return redirect("locations:mobile_import_list")

        except Exception as e:
            messages.error(request, f"Error processing request: {str(e)}")

    context = {
        "mobile_import": mobile_import,
        "import_data": json.dumps(mobile_import.import_data, indent=2),
    }

    return render(request, "locations/review_mobile_import.html", context)


@login_required
@role_required(["ADMIN"])
def bulk_import_actions(request):
    """Handle bulk actions on mobile imports"""
    if request.method == "POST":
        import_ids = request.POST.getlist("import_ids")
        action = request.POST.get("bulk_action")

        imports = MobileDataImport.objects.filter(id__in=import_ids)
        processed = 0

        for mobile_import in imports:
            try:
                if action == "approve":
                    mobile_import.approve(request.user)
                    processed += 1
                elif action == "reject":
                    mobile_import.reject(request.user, "Bulk rejected")
                    processed += 1
            except Exception as e:
                import logging

                logging.getLogger(__name__).error(
                    f"Failed to process import {mobile_import.id}: {e}"
                )
                continue

        messages.success(request, f"{processed} imports processed successfully.")

    return redirect("locations:mobile_import_list")
